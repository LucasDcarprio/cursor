"""
管理后台视图 - 仪表板、数据管理
"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from sqlalchemy import and_, or_
from datetime import datetime, timedelta
import os
import tempfile
import openpyxl

from app import db
from app.models.user import User
from app.models.store import Store
from app.models.product import Product
from app.models.order import Order
from app.views.auth import require_super, require_admin

admin_bp = Blueprint('admin', __name__)

@admin_bp.route('/')
@admin_bp.route('/dashboard')
@login_required
def dashboard():
    """管理后台首页"""
    # 获取统计数据
    stats = {}
    
    if current_user.is_super():
        # 超级管理员看到所有数据
        stats['total_users'] = User.query.count()
        stats['total_stores'] = Store.query.count()
        stats['total_products'] = Product.query.count()
        stats['total_orders'] = Order.query.count()
        stats['active_orders'] = Order.query.filter(Order.order_status.in_(['pending', 'processing', 'shipped'])).count()
    else:
        # 普通用户只看到自己的数据
        stats['total_stores'] = Store.query.filter_by(from_user=current_user.id).count()
        stats['total_products'] = Product.query.filter_by(from_user=current_user.id).count()
        stats['total_orders'] = Order.query.filter_by(from_user=current_user.id).count()
        stats['active_orders'] = Order.query.filter(
            and_(Order.from_user == current_user.id, 
                 Order.order_status.in_(['pending', 'processing', 'shipped']))
        ).count()
    
    # 获取最近订单
    recent_orders_query = Order.query
    if not current_user.is_super():
        recent_orders_query = recent_orders_query.filter_by(from_user=current_user.id)
    
    recent_orders = recent_orders_query.order_by(Order.created_at.desc()).limit(5).all()
    
    return render_template('admin_dashboard.html', stats=stats, recent_orders=recent_orders)

@admin_bp.route('/users')
@require_super
def users():
    """用户管理 - 仅超级管理员可访问"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = User.query
    if search:
        query = query.filter(
            or_(User.username.contains(search),
                User.phone.contains(search),
                User.brand_name.contains(search))
        )
    
    users = query.order_by(User.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('admin_users.html', users=users, search=search)

@admin_bp.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@require_super
def edit_user(user_id):
    """编辑用户 - 仅超级管理员可访问"""
    user = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        data = request.form
        
        # 更新用户信息
        user.role = data.get('role', user.role)
        user.max_stores = int(data.get('max_stores', user.max_stores))
        
        # 处理到期时间
        expire_time_str = data.get('expire_time')
        if expire_time_str:
            try:
                user.expire_time = datetime.strptime(expire_time_str, '%Y-%m-%d')
            except ValueError:
                flash('日期格式错误', 'error')
                return render_template('edit_user.html', user=user)
        
        try:
            db.session.commit()
            flash('用户信息更新成功', 'success')
            return redirect(url_for('admin.users'))
        except Exception as e:
            db.session.rollback()
            flash(f'更新失败：{str(e)}', 'error')
    
    return render_template('edit_user.html', user=user)

@admin_bp.route('/stores')
@login_required
def stores():
    """门店管理"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = Store.query
    if not current_user.is_super():
        # 普通用户只能看到自己的门店
        query = query.filter_by(from_user=current_user.id)
    
    if search:
        query = query.filter(
            or_(Store.store_name.contains(search),
                Store.store_id.contains(search),
                Store.address.contains(search))
        )
    
    stores = query.order_by(Store.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('admin_stores.html', stores=stores, search=search)

@admin_bp.route('/stores/add', methods=['GET', 'POST'])
@login_required
def add_store():
    """添加门店"""
    # 检查是否可以添加门店
    if not current_user.is_super() and not current_user.can_add_store():
        flash('已达到最大门店数量限制', 'error')
        return redirect(url_for('admin.stores'))
    
    if request.method == 'POST':
        data = request.form
        
        # 验证必填字段
        store_id = data.get('store_id', '').strip()
        store_name = data.get('store_name', '').strip()
        
        if not store_id or not store_name:
            flash('门店ID和名称为必填字段', 'error')
            return render_template('add_store.html')
        
        # 检查门店ID唯一性
        if Store.query.filter_by(store_id=store_id).first():
            flash('门店ID已存在', 'error')
            return render_template('add_store.html')
        
        try:
            # 创建新门店
            store = Store(
                store_id=store_id,
                store_name=store_name,
                address=data.get('address', ''),
                phone=data.get('phone', ''),
                mobile=data.get('mobile', ''),
                contact_name=data.get('contact_name', ''),
                province=data.get('province', ''),
                city=data.get('city', ''),
                district=data.get('district', ''),
                postcode=data.get('postcode', ''),
                district_code=data.get('district_code', ''),
                alibaba_openid=data.get('alibaba_openid', ''),
                alibaba_openkey=data.get('alibaba_openkey', ''),
                alibaba_token=data.get('alibaba_token', ''),
                from_user=current_user.id
            )
            
            db.session.add(store)
            db.session.commit()
            
            flash('门店添加成功', 'success')
            return redirect(url_for('admin.stores'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'添加失败：{str(e)}', 'error')
    
    return render_template('add_store.html')

@admin_bp.route('/data')
@login_required
def data():
    """基础数据管理"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    
    query = Product.query
    if not current_user.is_super():
        # 普通用户只能看到自己的数据
        query = query.filter_by(from_user=current_user.id)
    
    if search:
        query = query.filter(
            or_(Product.product_name.contains(search),
                Product.sku_code.contains(search),
                Product.manufacturer.contains(search))
        )
    
    products = query.order_by(Product.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    return render_template('admin_data.html', products=products, search=search)

@admin_bp.route('/data2')
@login_required
def data2():
    """采购订单管理"""
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '', type=str)
    status = request.args.get('status', '', type=str)
    
    query = Order.query
    if not current_user.is_super():
        # 普通用户只能看到自己的订单
        query = query.filter_by(from_user=current_user.id)
    
    if search:
        query = query.filter(
            or_(Order.product_name.contains(search),
                Order.order_number.contains(search),
                Order.sku_code.contains(search))
        )
    
    if status:
        query = query.filter_by(order_status=status)
    
    orders = query.order_by(Order.created_at.desc()).paginate(
        page=page, per_page=20, error_out=False
    )
    
    # 获取状态列表用于筛选
    status_options = [
        ('pending', '待处理'),
        ('processing', '处理中'),
        ('shipped', '已发货'),
        ('delivered', '已送达'),
        ('completed', '已完成'),
        ('cancelled', '已取消'),
        ('failed', '失败')
    ]
    
    return render_template('admin_data2.html', orders=orders, search=search, 
                         status=status, status_options=status_options)

@admin_bp.route('/export/<data_type>')
@login_required
def export_data(data_type):
    """导出数据为Excel"""
    try:
        # 创建临时文件
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        
        if data_type == 'products':
            # 导出商品数据
            query = Product.query
            if not current_user.is_super():
                query = query.filter_by(from_user=current_user.id)
            
            products = query.all()
            data = [product.to_dict() for product in products]
            filename = f'商品数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
        elif data_type == 'orders':
            # 导出订单数据
            query = Order.query
            if not current_user.is_super():
                query = query.filter_by(from_user=current_user.id)
            
            orders = query.all()
            data = [order.to_dict() for order in orders]
            filename = f'订单数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
        elif data_type == 'stores':
            # 导出门店数据
            query = Store.query
            if not current_user.is_super():
                query = query.filter_by(from_user=current_user.id)
            
            stores = query.all()
            data = [store.to_dict() for store in stores]
            filename = f'门店数据_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
        else:
            flash('不支持的数据类型', 'error')
            return redirect(url_for('admin.dashboard'))
        
        # 写入Excel文件（使用openpyxl替代pandas）
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if data:
            # 写入表头
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=item.get(header, ''))
        
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(temp_file.name, as_attachment=True, download_name=filename)
        
    except Exception as e:
        flash(f'导出失败：{str(e)}', 'error')
        return redirect(url_for('admin.dashboard'))

@admin_bp.route('/import/<data_type>', methods=['POST'])
@login_required
def import_data(data_type):
    """导入Excel数据"""
    if 'file' not in request.files:
        flash('请选择文件', 'error')
        return redirect(request.referrer)
    
    file = request.files['file']
    if file.filename == '':
        flash('请选择文件', 'error')
        return redirect(request.referrer)
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        flash('请上传Excel文件', 'error')
        return redirect(request.referrer)
    
    try:
        # 读取Excel文件
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # 获取表头和数据
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # 跳过空行
                row_dict = dict(zip(headers, row))
                data.append(row_dict)
        
        success_count = 0
        error_count = 0
        
        if data_type == 'products':
            # 导入商品数据
            for row in data:
                try:
                    product = Product(
                        barcode=str(row.get('barcode', '') or ''),
                        sku_code=str(row.get('sku_code', '') or ''),
                        product_name=str(row.get('product_name', '') or ''),
                        specification=str(row.get('specification', '') or ''),
                        purchase_attributes=str(row.get('purchase_attributes', '') or ''),
                        settlement_link=str(row.get('settlement_link', '') or ''),
                        unit_count=str(row.get('unit_count', '') or ''),
                        min_order=str(row.get('min_order', '') or ''),
                        meituan_price=str(row.get('meituan_price', '') or ''),
                        manufacturer=str(row.get('manufacturer', '') or ''),
                        from_user=current_user.id
                    )
                    db.session.add(product)
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    print(f"导入商品失败: {e}")
        
        elif data_type == 'orders':
            # 导入订单数据
            for row in data:
                try:
                    order = Order(
                        store_id=str(row.get('store_id', '') or ''),
                        product_name=str(row.get('product_name', '') or ''),
                        sku_code=str(row.get('sku_code', '') or ''),
                        purchase_quantity=str(row.get('purchase_quantity', '') or ''),
                        purchase_price=str(row.get('purchase_price', '') or ''),
                        supplier=str(row.get('supplier', '') or ''),
                        order_status=str(row.get('order_status', '') or 'pending'),
                        from_user=current_user.id
                    )
                    db.session.add(order)
                    success_count += 1
                except Exception as e:
                    error_count += 1
                    print(f"导入订单失败: {e}")
        
        db.session.commit()
        flash(f'导入完成：成功 {success_count} 条，失败 {error_count} 条', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'导入失败：{str(e)}', 'error')
    
    return redirect(request.referrer)