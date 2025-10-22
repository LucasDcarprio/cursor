"""
阿里巴巴 1688 平台 API 集成模块
Alibaba 1688 Platform API Integration Module

提供与阿里巴巴 1688 平台的完整 API 集成功能，包括：
- 商品信息查询
- 订单创建和管理
- 支付处理
- 物流跟踪
- 优惠券管理

作者：泓枢智创团队
版本：1.0.0
"""

import requests
import time
import json
import hmac
import hashlib
import urllib.parse
import config
import openpyxl
import re
import concurrent.futures
import os
import logging
import csv
import sqlite3
from flask import session, has_request_context
from functools import wraps
from datetime import datetime
# 从数据库获取当前登录用户的API凭证
def get_current_user_credentials(user_id=None, store_id=None):
    try:
        # 如果未提供user_id，尝试从session获取
        if user_id is None and has_request_context() and 'user_id' in session:
            user_id = session['user_id']
        
        # 连接数据库
        conn = sqlite3.connect('./instance/data.db')
        cursor = conn.cursor()
        # 如果提供了store_id，优先从store表获取凭证
        if store_id:
            cursor.execute("""
                SELECT alibaba_openid, alibaba_openkey, alibaba_token 
                FROM store 
                WHERE store_id = ?
            """, (store_id,))
            store_data = cursor.fetchone()
            print(f'获取到门店信息{store_data}')
            if store_data and all(store_data):
                conn.close()
                return store_data[0], store_data[1], store_data[2]
            else:
                print(f"未能获取{store_id}的阿里巴巴有效凭证")
                return ((f"未能获取{store_id}的阿里巴巴有效凭证"))
        
        # 如果以上都未获取到有效凭证，使用默认值
        print(f"未能获取有效凭证，使用默认凭证")
        return config.API_KEY, config.API_SECRET, config.API_TOKEN
    except Exception as e:
        print(f"获取凭证时出错: {e}")
        # 出错时使用默认值
        return config.API_KEY, config.API_SECRET, config.API_TOKEN

# API配置（初始值，将在每次API调用前从数据库重新加载）
API_KEY, API_SECRET, API_TOKEN = get_current_user_credentials()
BASE_URL = 'https://gw.open.1688.com/openapi/'
# API路径
API_PATHS = config.API_PATHS
ADDRESS = config.ADDRESS
PARAMAS = {
    "_aop_timestamp": str(int(time.time() * 1000)),
    "access_token": API_TOKEN,
    "domain": "junmanbaihuo.1688.com",
    #获取商品信息
    "productID":619977523250,
    "webSite":"1688",
    #免密支付
    "tradeWithholdPreparePayParam": json.dumps(
        {
            "orderId":2212893950301027671
        },
    ),
    #应用级输入参数
    "flow": "general",  # 创建大市场订单
    "message": "阿里巴巴后台测试订单",  # 买家留言（可选）
    #收货地址信息
    "addressParam": json.dumps(ADDRESS),
    #下单数据
    "cargoParamList": json.dumps([
        {
            "specId": "e44bfa7c69c84ca501d30dae2858d6ae",
            "quantity": 10,
            "offerId": 619977523250
        },
    ]),
    "fenxiaoChannel": "SKU-ID",  # 回流订单下游平台（可选）
    "useRedEnvelope": "y"
}

# 添加API重试装饰器
def api_retry(max_retries=2, retry_delay=0.1):
    """
    API调用重试装饰器
    
    Args:
        max_retries: 最大重试次数
        retry_delay: 重试间隔（秒）
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            retries = 0
            last_error = None
            
            # 从kwargs中获取user_id，如果存在
            user_id = kwargs.get('user_id')
            
            while retries < max_retries:
                try:
                    result = func(*args, **kwargs)
                    
                    # 检查是否是错误响应
                    if isinstance(result, dict) and result.get("error") == True:
                        error_msg = result.get("error_message", "未知API错误")
                        retries += 1
                        
                        # 特定错误可能需要特殊处理
                        if "签名无效" in error_msg or "access_token无效" in error_msg:
                            print(f"API凭证错误，尝试从数据库重新获取用户凭证...")
                            # 直接从数据库重新获取用户凭证
                            global API_KEY, API_SECRET, API_TOKEN
                            API_KEY, API_SECRET, API_TOKEN = get_current_user_credentials(user_id)
                        
                        print(f"API调用失败: {error_msg}, 第{retries}次重试 (最大{max_retries}次)")
                        time.sleep(retry_delay)
                        continue
                    
                    return result
                    
                except (requests.RequestException, ConnectionError, TimeoutError) as e:
                    retries += 1
                    last_error = str(e)
                    print(f"网络错误: {last_error}, 第{retries}次重试 (最大{max_retries}次)")
                    time.sleep(retry_delay)
                except Exception as e:
                    # 对于其他类型错误，直接返回错误结果不重试
                    return {"error": True, "error_message": f"异常错误: {str(e)}"}
            
            # 达到最大重试次数后返回最后的错误
            return {"error": True, "error_message": f"达到最大重试次数({max_retries}): {last_error}"}
        
        return wrapper
    return decorator

def log_sku_info(product_id, sku_infos):
    return
    # 确保logs文件夹存在
    os.makedirs('logs', exist_ok=True)
    
    # 设置日志文件名和格式
    log_filename = f"logs/sku_info_log_.log"
    logging.basicConfig(
        filename=log_filename,
        level=logging.INFO,
        format='%(asctime)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        encoding='utf-8'
    )

    # 准备并记录日志
    log_entry = {
        "product_id": product_id,
        "sku_infos": sku_infos
    }
    logging.info(json.dumps(log_entry, ensure_ascii=False))

def generate_signature(url_path, params, secret_key):
    sorted_params = sorted(params.items(), key=lambda x: x[0])
    params_str = ''.join(f"{k}{v}" for k, v in sorted_params)
    sign_factor = f"{url_path}{params_str}"
    
    # print(',代签名字符串==', sign_factor)
    hmac_code = hmac.new(secret_key.encode(), sign_factor.encode(), hashlib.sha1)
    return hmac_code.hexdigest().upper()

def build_url(path, params, user_id=None, store_id=None):
    # 检查API_KEY是否有效
    if not API_KEY or API_KEY == "0":
        print("警告：API_KEY无效，尝试重新获取用户凭证")
        # 再次尝试获取用户凭证
        api_key, _, _ = get_current_user_credentials(user_id, store_id)
        if not api_key or api_key == "0":
            print("仍然无法获取有效的API_KEY，可能会导致请求失败")
            api_key = "0"  # 使用默认值
    else:
        api_key = API_KEY
        
    full_path = f'param2/1/{path}/{api_key}'
    query_string = "&".join(f"{k}={urllib.parse.quote(str(v), safe='')}" for k, v in params.items())
    url = f"{BASE_URL}{full_path}?{query_string}"
    
    # 使用当前用户的API_SECRET生成签名
    api_secret = API_SECRET
    params['_aop_signature'] = generate_signature(full_path, params, api_secret)
    
    return f"{url}&_aop_signature={params['_aop_signature']}"


def make_api_request(path, params, user_id=None, store_id=None):
    # 从数据库获取最新的用户API凭证
    global API_KEY, API_SECRET, API_TOKEN
    try:
        # 直接从数据库获取指定用户或门店的凭证
        API_KEY, API_SECRET, API_TOKEN = get_current_user_credentials(user_id, store_id)
        
        # 确保access_token使用最新的token
        if "access_token" in params:
            params["access_token"] = API_TOKEN
        
        url = build_url(path, params, user_id, store_id)
        response = requests.get(url, timeout=30)  # 添加超时设置
        
        # 检查HTTP响应状态
        if response.status_code != 200:
            print(f"API请求失败: HTTP状态码 {response.status_code}")
            print(f"响应内容: {response.text[:200]}")
            return {"error": True, "error_message": f"HTTP错误: {response.status_code}", "response": response.text[:200]}
        
        # 尝试解析JSON响应
        try:
            result = response.json()
            
            # 检查API错误
            if "error_message" in result or "errorMessage" in result:
                error_msg = result.get("error_message") or result.get("errorMessage")
                print(f"API返回错误: {error_msg}")
                return {"error": True, "error_message": error_msg}
            
            return result
        except json.JSONDecodeError as e:
            print(f"解析JSON响应失败: {e}")
            print(f"响应内容: {response.text[:200]}")
            return {"error": True, "error_message": f"JSON解析错误: {str(e)}", "response": response.text[:200]}
            
    except Exception as e:
        print(f"API请求发生异常: {str(e)}")
        return {"error": True, "error_message": f"请求异常: {str(e)}"}

def extract_product_id(url):
    if 'yangkeduo' in url:
        print('拼多多商品！')
        return False
    match = re.search(r'offer/(\d+)\.html', url)
    return int(match.group(1)) if match else None


def get_address(user_id=None, store_id=None):
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
    }
    result = make_api_request(API_PATHS['get_address'], params, user_id=user_id, store_id=store_id)
    # print(result)
    # print(result['result']['receiveAddressItems'][0]['id'])
    return [result['result']['receiveAddressItems'][0]['id'],result['result']['receiveAddressItems']]#返回默认地址

def get_skuinfo(productID, user_id=None, store_id=None):
    """
    根据商品ID获取商品的SKU信息
    
    Args:
        productID: 商品ID
        user_id: 用户ID（可选）
        store_id: 商店ID（可选）
        
    Returns:
        成功时返回SKU信息列表，失败时返回空列表或错误信息
    """
    # 构建API请求参数
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "productID": int(productID),
        "webSite": "1688"  # 指定站点为1688
    }
    
    # 发送API请求获取商品信息
    result = make_api_request(API_PATHS['get_goods_info'], params, user_id=user_id, store_id=store_id)
    
    # 检查是否有错误信息
    if isinstance(result, dict) and result.get('error') == True:
        print(f"Error: {result.get('error_message', '未知错误')}")
        return []
    
    if 'errMsg' in result and result['errMsg']:
        print(f"API返回错误: {result['errMsg']}")
        return []
    
    # 获取商品信息
    product_info = result.get('productInfo', {})
    if not product_info:
        print(f"未获取到商品ID {productID} 的信息")
        return []
    
    # 从productInfo中获取skuInfos列表
    sku_infos = product_info.get('skuInfos', [])
    
    # 记录SKU信息（用于日志）
    log_sku_info(productID, sku_infos)
    
    # 如果需要调试，可以打印SKU信息
    if sku_infos:
        print(f"成功获取商品ID {productID} 的SKU信息，共 {len(sku_infos)} 个SKU")
    else:
        print(f"商品ID {productID} 没有SKU信息或获取失败")
    
    return sku_infos

def get_skuinfo_detail(productID, user_id=None, store_id=None):
    """
    根据商品ID获取商品的详细SKU信息，包括价格、库存、属性等
    
    Args:
        productID: 商品ID
        user_id: 用户ID（可选）
        store_id: 商店ID（可选）
        
    Returns:
        成功时返回详细的SKU信息列表，失败时返回空列表
    """
    # 构建API请求参数
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "productID": int(productID),
        "webSite": "1688"  # 指定站点为1688
    }
    
    # 发送API请求获取商品信息
    result = make_api_request(API_PATHS['get_goods_info'], params, user_id=user_id, store_id=store_id)
    
    # 检查是否有错误信息
    if isinstance(result, dict) and result.get('error') == True:
        print(f"Error: {result.get('error_message', '未知错误')}")
        return []
        
    if 'errMsg' in result and result['errMsg']:
        print(f"API返回错误: {result['errMsg']}")
        return []
    
    # 获取商品信息
    product_info = result.get('productInfo', {})
    if not product_info:
        print(f"未获取到商品ID {productID} 的信息")
        return []
    
    # 从productInfo中获取skuInfos列表
    sku_infos = product_info.get('skuInfos', [])
    
    # 记录SKU信息（用于日志）
    log_sku_info(productID, sku_infos)
    
    if not sku_infos:
        print(f"商品ID {productID} 没有SKU信息或获取失败")
        return []
        
    print(f"成功获取商品ID {productID} 的SKU信息，共 {len(sku_infos)} 个SKU")
    
    # 处理SKU详细信息
    sku_details = []
    for i, sku in enumerate(sku_infos, 1):
        try:
            sku_detail = {
                "sku_number": i,
                "specId": sku.get('specId'),
                "skuId": sku.get('skuId'),
                "price": sku.get('price'),
                "amountOnSale": sku.get('amountOnSale'),
                "attributes": {}
            }
            
            # 处理属性信息
            attributes = sku.get('attributes', [])
            for attr in attributes:
                attr_name = attr.get('attributeName')
                attr_value = attr.get('attributeValue')
                if attr_name and attr_value:
                    sku_detail['attributes'][attr_name] = attr_value
            
            sku_details.append(sku_detail)
        except Exception as e:
            print(f"处理SKU详情时出错: {str(e)}")
            continue
    
    return sku_details

def find_matching_spec_id(sku_infos, target_specs):
    """
    在SKU信息列表中查找匹配的规格ID
    
    Args:
        sku_infos: SKU信息列表
        target_specs: 目标规格字符串（格式如 "红色,XL" 或 "红色;XL"）
        
    Returns:
        匹配的specId，如果没有匹配则返回None
    """
    if not sku_infos:
        print("SKU信息列表为空，无法匹配规格")
        return None
        
    # 确保target_specs是字符串类型
    if target_specs is None:
        print("目标规格为空，无法匹配")
        return None
        
    target_specs = str(target_specs)
    # 处理不同的分隔符（逗号、分号）
    target_spec_list = re.split(r'[;,]', target_specs)
    # 去除空白字符
    target_spec_list = [spec.strip() for spec in target_spec_list if spec.strip()]
    
    if not target_spec_list:
        print("处理后的目标规格列表为空，无法匹配")
        return None
    
    print(f"正在匹配规格: {target_spec_list}")
    
    # 遍历SKU列表，查找匹配项
    for sku in sku_infos:
        try:
            attributes = sku.get('attributes', [])
            if not attributes:
                continue
                
            # 获取属性值列表
            attribute_values = [str(attr.get('attributeValue', '')).strip() for attr in attributes]
            # 输出调试信息
            # print(f"SKU属性值: {attribute_values}")
            
            # 检查所有目标规格是否都在当前SKU的属性值中
            if all(any(target_spec.lower() in attr_val.lower() for attr_val in attribute_values) 
                  for target_spec in target_spec_list):
                spec_id = sku.get('specId')
                if spec_id:
                    print(f'{target_specs} -> 成功匹配specID：{spec_id}')
                    return spec_id
        except Exception as e:
            print(f"匹配规格时出错: {str(e)}")
            continue
    
    print(f"未找到匹配规格 '{target_specs}' 的specID")
    return None

def process_row(row, sheet):
    url = sheet.cell(row=row, column=6).value  # F列
    target_specs = sheet.cell(row=row, column=7).value  # G列
    
    if url and target_specs:
        product_id = extract_product_id(url)
        
        if product_id:
            sku_infos = get_skuinfo(product_id)
            # 检查返回的sku_infos是否为空列表
            if not sku_infos:
                print(f"警告: 商品ID {product_id} 在第 {row} 行没有SKU信息")
                return row, product_id, None
            
            matching_spec_id = find_matching_spec_id(sku_infos, target_specs)
            return row, product_id, matching_spec_id
    
    return row, None, None



def process_excel():
    # 打开现有Excel文件
    wb = openpyxl.load_workbook('基础数据库2.xlsx')
    sheet = wb['新品']

    # 提示用户输入起始行和结束行
    start_row = 6267
    end_row = 7216

    # 确保输入的行数在有效范围内
    max_row = sheet.max_row
    start_row = max(2, start_row)  # 最小从第2行开始读取，因为第1行通常是标题
    end_row = min(max_row, end_row)

    # 创建一个CSV文件用于保存处理后的数据
    with open('基础数据库2_3processed.csv', mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)

        # 写入标题行
        headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
        headers.extend(['Product ID', 'Matching Spec ID'])
        writer.writerow(headers)

        # 遍历行并处理数据
        for row in range(start_row, end_row + 1):
            row_num, product_id, matching_spec_id = process_row(row, sheet)
            
            # 复制原数据到新文件
            row_data = [sheet.cell(row=row_num, column=col).value for col in range(1, sheet.max_column + 1)]
            row_data.extend([product_id, matching_spec_id])
            writer.writerow(row_data)

            # 打印调试信息
            # print(f"Processed row {row_num}: Product ID = {product_id}, Matching Spec ID = {matching_spec_id}")

    print("CSV processing completed and saved to 基础数据库2_processed.csv")

#get_logistics_info获取物流信息
def get_wuliu(order_id, user_id=None, store_id=None):
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "orderId": order_id,
        "webSite": '1688',
    }
    result = make_api_request(API_PATHS['get_logistics_info'], params, user_id=user_id, store_id=store_id)
    # return result
    if 'result' not in result:
        print(f"获取订单{order_id}的物流信息失败: {result.get('errorMessage', '未知错误')}")
        return False
    
    logistics_bill_no = result['result'][0].get('logisticsBillNo', '')
    if logistics_bill_no:
        print(f"订单{order_id}成功匹配物流单号：{logistics_bill_no}")
        return {
            'logistics_bill_no': logistics_bill_no,
            'status': result.get('status', ''),
            'logistics_company': result.get('logisticsCompanyName', '')
        }
    else:
        print(f"未找到订单{order_id}的物流单号{result}")
        return False

#下单前预览订单
def order_preview(goods_id, specId, number, user_id=None, store_id=None):
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "domain": "junmanbaihuo.1688.com",
        "flow": "general",  # 创建大市场订单
        "message": "阿里巴巴后台订单,企业合作,请减免运费！",  # 买家留言（可选）
        #收货地址信息
        "addressParam":{'addressId':3173448020},
        #下单数据
        "cargoParamList": json.dumps([
            {
                "specId": specId,
                "quantity": number,
                "offerId": goods_id,
            },
        ]),
        "fenxiaoChannel": "",  # 回流订单下游平台（可选）
        "useRedEnvelope": "y"
    }
    result = make_api_request(API_PATHS['order_preview'], params, user_id=user_id, store_id=store_id)
    # print(result)
    return result
    
def after_order_preview(order_id, query_type=None, user_id=None, store_id=None):
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "webSite": "1688",
        "orderId": order_id,
        # "orderStatus":"waitbuyerpay",#状态未付款
    }
    result = make_api_request(API_PATHS['after_order_preview'], params, user_id=user_id, store_id=store_id)
    # print(result)
    if 'error_message' in result:
        print(result)
        return False
    if query_type == 'totalAmount':
        return float(result['result']['baseInfo'].get('totalAmount', 0))
        # print(result['result']['baseInfo']['totalAmount'])  # 总付款额 status
    if result['result']['baseInfo']['status'] == 'cancel' :
        print(result['result']['baseInfo']['status'])
        return '订单关闭'
    if result['result']['baseInfo']['status'] == 'waitsellersend' :
        print(result['result']['baseInfo']['status'])
        return '已支付'
    shippingFee = float(result['result']['baseInfo']['shippingFee'])
    # print(shippingFee)  # 运费
    # print(result['result']['productItems'])  # 规格清单信息
    # 获取折扣费用
    discountFee = 0
    if 'newStepOrderList' in result['result']['baseInfo'] and len(result['result']['baseInfo']['newStepOrderList']) > 0:
        discountFee = float(result['result']['baseInfo']['newStepOrderList'][0]['discountFee'])
        print(f'获取到优惠金额：{discountFee}')
    
    all_quantity = 0
    product_details = []  # 用于存储每个商品的最终单价和 SKU 信息

    for item in result['result']['productItems']:
        # print(f'当前item==> \n {item}')
        productID = item['productID']
        # print(f'商品单价: {item["price"]}')
        item_price = float(item['price'])
        # print(f'商品规格ID: {item["specId"]}')
        # 只获取 skuInfos 中的 value 部分
        sku_infos_values = item.get('specId', [])
        # print(f'单个商品数量: {item["quantity"]}')
        all_quantity += int(item['quantity'])  # 总采购数量 = 单品数量之和
        
        # 计算商品的最终单价
        product_details.append({
            'final_price': item_price,  # 先不加运费，后面统一加
            'sku_infos_values': sku_infos_values,
            'quantity': int(item['quantity']),
            'productId':productID
        })
        
    #在运费里面减去优惠金额，这里为了方便，其实也可以在单个商品里减去
    shippingFee = shippingFee - discountFee if discountFee > 0 else shippingFee
    # 计算平均运费单价
    per_shippingFee = shippingFee / all_quantity

    # print(f'平均运费单价：{per_shippingFee}')

    # 更新每个商品的最终单价，加上平均运费单价
    for product in product_details:
        product['final_price'] += per_shippingFee
    print(product_details)
    return product_details

def get_order_list(start_time=None, end_time=None, status=None, page=1, page_size=50, user_id=None, store_id=None):
    """
    获取阿里巴巴订单列表，支持按时间范围和订单状态筛选，自动翻页获取全部结果
    
    Args:
        start_time: 开始时间（格式：YYYY-MM-DD）
        end_time: 结束时间（格式：YYYY-MM-DD）
        status: 订单状态（waitbuyerpay-等待买家付款, waitsellersend-等待卖家发货, waitbuyerreceive-等待买家收货等）
            如果为空字符串或None，则查询所有状态
        page: 页码，默认为1
        page_size: 每页大小，默认为50
        user_id: 用户ID（可选）
        store_id: 商店ID（可选）
        
    Returns:
        订单详细信息列表
    """
    # 构建API请求参数的基础结构
    base_params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "pageSize": page_size,  # 限制最大为50，避免超出API限制
        "needBuyerAddressAndPhone": "true",  # 获取买家地址和电话
        "needMemoInfo": "true",  # 获取备注信息
        "webSite": "1688"  # 指定站点为1688
    }
    
    # 添加可选的筛选参数（只在status有值且不为空字符串时添加orderStatus参数）
    if status and status.strip():
        base_params["orderStatus"] = status
    
    # 处理日期范围
    if start_time:
        try:
            # 将YYYY-MM-DD转换为API所需的日期格式
            start_date = datetime.strptime(start_time, "%Y-%m-%d")
            # 格式化为API要求的格式：yyyyMMddHHmmssSSS+0800
            base_params["createStartTime"] = start_date.strftime("%Y%m%d%H%M%S") + "000+0800"
        except ValueError:
            print(f"无效的开始日期格式: {start_time}，应为YYYY-MM-DD")
    
    if end_time:
        try:
            # 将YYYY-MM-DD转换为API所需的日期格式
            end_date = datetime.strptime(end_time, "%Y-%m-%d")
            # 设置为当天的23:59:59
            end_date = end_date.replace(hour=23, minute=59, second=59)
            # 格式化为API要求的格式：yyyyMMddHHmmssSSS+0800
            base_params["createEndTime"] = end_date.strftime("%Y%m%d%H%M%S") + "000+0800"
        except ValueError:
            print(f"无效的结束日期格式: {end_time}，应为YYYY-MM-DD")
    
    # 存储所有处理后的订单数据
    all_processed_orders = []
    current_page = page
    total_pages = None
    empty_page_count = 0  # 计数连续空页面
    max_empty_pages = 3   # 连续多少个空页面后停止
    
    print(f"开始查询订单数据，每页 {page_size} 条，将获取全部数据...")
    
    # 循环获取所有页的数据
    while True:
        # 添加当前页码到参数
        params = base_params.copy()
        params["page"] = current_page
        
        print(f"正在获取第 {current_page} 页数据...")
        # 发送API请求
        result = make_api_request(API_PATHS['get_order_list'], params, user_id=user_id, store_id=store_id)
        
        # 检查是否有错误
        if isinstance(result, dict) and result.get('error') == True:
            print(f"获取订单列表第{current_page}页失败: {result.get('error_message', '未知错误')}")
            break
        
        # 计算总页数（仅第一次请求时）
        if total_pages is None and 'totalRecord' in result:
            total_records = int(result.get('totalRecord', 0))
            total_pages = (total_records + page_size - 1) // page_size
            print(f"总记录数: {total_records}, 总页数: {total_pages}")
        
        # 处理订单数据，提取关键信息
        page_has_data = False
        if 'result' in result and isinstance(result['result'], list) and result['result']:
            page_has_data = True
            empty_page_count = 0  # 重置空页面计数
            
            for order in result['result']:
                try:
                    base_info = order.get('baseInfo', {})
                    
                    # 提取买家和卖家联系信息
                    buyer_contact = base_info.get('buyerContact', {})
                    seller_contact = base_info.get('sellerContact', {})
                    
                    # 提取收货人信息
                    receiver_info = base_info.get('receiverInfo', {})
                    
                    # 提取订单基本信息
                    order_data = {
                        'orderId': base_info.get('id', ''),
                        'createTime': base_info.get('createTime', ''),
                        'modifyTime': base_info.get('modifyTime', ''),
                        'payTime': '',  # 从交易条款中获取
                        'completeTime': base_info.get('completeTime', ''),
                        'totalAmount': base_info.get('totalAmount', 0),
                        'shippingFee': base_info.get('shippingFee', 0),
                        'discount': base_info.get('discount', 0),
                        'status': base_info.get('status', ''),
                        'statusStr': '',  # 后面根据状态码转换
                        'tradeType': base_info.get('tradeType', ''),
                        'buyerLoginId': base_info.get('buyerLoginId', ''),
                        'sellerLoginId': base_info.get('sellerLoginId', ''),
                        'buyerCompanyName': buyer_contact.get('companyName', ''),
                        'sellerCompanyName': seller_contact.get('companyName', ''),
                        'buyerContact': buyer_contact.get('name', ''),
                        'buyerPhone': buyer_contact.get('phone', ''),
                        'sellerContact': seller_contact.get('name', ''),
                        'sellerPhone': seller_contact.get('phone', ''),
                        'receiverName': receiver_info.get('toFullName', ''),
                        'receiverArea': receiver_info.get('toArea', ''),
                        'closeReason': base_info.get('closeReason', ''),
                        'refund': base_info.get('refund', 0),
                        'products': []
                    }
                    
                    # 根据状态码填充状态描述
                    status_map = {
                        'waitbuyerpay': '等待买家付款',
                        'waitsellersend': '等待卖家发货',
                        'waitbuyerreceive': '等待买家收货',
                        'success': '交易成功',
                        'cancel': '交易取消',
                        'terminated': '交易终止'
                    }
                    order_data['statusStr'] = status_map.get(order_data['status'], order_data['status'])
                    
                    # 从交易条款中获取支付时间等信息
                    if 'tradeTerms' in order and isinstance(order['tradeTerms'], list):
                        for term in order['tradeTerms']:
                            if 'payTime' in term:
                                order_data['payTime'] = term.get('payTime', '')
                    
                    # 提取产品信息
                    for product in order.get('productItems', []):
                        product_data = {
                            'productId': product.get('productID', ''),
                            'subItemID': product.get('subItemID', ''),
                            'name': product.get('name', ''),
                            'price': product.get('price', 0),
                            'quantity': product.get('quantity', 0),
                            'itemAmount': product.get('itemAmount', 0),
                            'skuId': product.get('skuID', ''),
                            'specId': product.get('specId', ''),
                            'unit': product.get('unit', ''),
                            'weight': product.get('weight', ''),
                            'productImgUrl': product.get('productImgUrl', []),
                            'productSnapshotUrl': product.get('productSnapshotUrl', ''),
                            'status': product.get('status', ''),
                            'statusStr': product.get('statusStr', ''),
                            'logisticsStatus': product.get('logisticsStatus', ''),
                            'entryDiscount': product.get('entryDiscount', 0),
                        }
                        
                        # 添加SKU信息
                        if 'skuInfos' in product:
                            product_data['skuInfos'] = product.get('skuInfos', [])
                        
                        order_data['products'].append(product_data)
                    
                    all_processed_orders.append(order_data)
                except Exception as e:
                    print(f"处理订单数据时出错: {str(e)}")
                    continue
        else:
            # 页面没有数据
            empty_page_count += 1
            print(f"第 {current_page} 页没有订单数据，连续空页面数: {empty_page_count}")
            
            # 如果连续多个页面都是空的，认为已经获取完毕
            if empty_page_count >= max_empty_pages:
                print(f"连续 {max_empty_pages} 页没有数据，停止获取")
                break
        
        # 定期显示处理进度
        if current_page % 10 == 0:
            print(f"已处理 {current_page} 页，获取到 {len(all_processed_orders)} 条记录")
        
        # 如果已经到达服务器返回的总页数，跳出循环
        if total_pages is not None and current_page >= total_pages:
            print(f"已到达最后一页 {current_page}/{total_pages}，停止获取")
            break
            
        # 增加页码，继续查询下一页
        current_page += 1
        
        # 避免请求过快，增加小延迟
        time.sleep(0.2)
    
    print(f"完成！总共获取了 {len(all_processed_orders)} 条订单数据")
    
    # 返回处理后的数据和分页信息
    return {
        'orders': all_processed_orders,
        'pagination': {
            'currentPage': page,
            'pageSize': page_size,
            'totalPages': total_pages or 0,
            'totalRecords': len(all_processed_orders)
        }
    }


def pay_without_pwd(order_id, user_id=None, store_id=None):
    params = {
    "_aop_timestamp": str(int(time.time() * 1000)),
    "access_token": API_TOKEN,
    #免密支付
    "tradeWithholdPreparePayParam": json.dumps(
        {
            "orderId":order_id
        },
    ),
    }
    result = make_api_request(API_PATHS['pay_without_password'], params, user_id=user_id, store_id=store_id)
    # print(result)
    return result

def make_order_paired(items, addressId, user_id=None, store_id=None):
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "domain": "junmanbaihuo.1688.com",
        "flow": "paired",
        "message": "阿里巴巴大客户后台订单,请优惠一些！降低运费可增加复购率，请务必尽快发货并上传运单编号,否则将会被系统自动退单！",
        "addressParam": {'addressId':addressId},
        "cargoParamList": json.dumps(items),
        "outOrderId": f"nn_{int(time.time())}",
        "useRedEnvelope": "y"
    }
    
    result = make_api_request(API_PATHS['make_order'], params, user_id=user_id, store_id=store_id)
    print("Paired Flow API Response:", result)
    
    if isinstance(result, dict) and result.get('success') == True:
        order_id = result['result']['orderId']
        total_amount = result['result']['totalSuccessAmount']
        post_fee = result['result']['postFee']
        print(f"Paired Flow Order placed successfully. Order ID: {order_id}, Total Amount: {total_amount}, Postage Fee: {post_fee}")
        return f"success:{order_id}"
    return None

def make_order_general(items, addressId, user_id=None, store_id=None):
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "domain": "junmanbaihuo.1688.com",
        "flow": "general",
        "message": "阿里巴巴大客户后台订单,请优惠一些！降低运费可增加复购率，请务必尽快发货并上传运单编号,否则将会被系统自动退单！",
        "addressParam": {'addressId':addressId},
        "cargoParamList": json.dumps(items),
        "outOrderId": f"nn_{int(time.time())}",
        "useRedEnvelope": "y"
    }
    
    result = make_api_request(API_PATHS['make_order'], params, user_id=user_id, store_id=store_id)
    print("General Flow API Response:", result)
    
    if isinstance(result, dict) and result.get('success') == True:
        order_id = result['result']['orderId']
        total_amount = result['result']['totalSuccessAmount']
        post_fee = result['result']['postFee']
        print(f"General Flow Order placed successfully. Order ID: {order_id}, Total Amount: {total_amount}, Postage Fee: {post_fee}")
        return f"success:{order_id}"
    elif isinstance(result, dict):
        error_message = result.get('message', '未知错误')
        return f"下单失败: {error_message}"
    return f"意外的响应类型: {type(result)}"

def make_order(items, user_id=None, store_id=None):
    addressId = get_address(user_id, store_id)
    addressId = int(addressId[0]) # 获取默认地址
    print(f'正在为{store_id}下单')
    
    try:
        # 首先尝试使用 paired 流程下单
        result = make_order_paired(items, addressId, user_id, store_id)
        
        # 如果 paired 流程失败，尝试使用 general 流程
        if not result:
            print("Paired flow failed, trying general flow...")
            result = make_order_general(items, addressId, user_id, store_id)
        
        return result

    except Exception as e:
        print(f"Order failed: {str(e)}")
        return f"下单异常: {str(e)}"
        
#批量下单
from query_data import *
import math
def validate_and_convert_to_int(value, default=0):
    """
    验证并将输入值转换为整数。
    如果值为None、NaN或无效，则返回默认值。
    """
    try:
        # 检查值是否为None或NaN
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return default
        
        # 尝试将值转换为浮点数再转换为整数
        return int(float(value))
    except (ValueError, TypeError):
        # 如果转换失败，返回默认值
        return default
        
@api_retry(max_retries=2, retry_delay=0.5)
def claim_optimal_coupon(offer_ids, user_id=None, store_id=None):
    """
    通过商品领取最优化的优惠券
    
    Args:
        offer_ids: 商品ID列表
        user_id: 用户ID（可选）
        store_id: 商店ID（可选）
        
    Returns:
        成功时返回API响应结果，失败时返回错误信息
    """
    if not offer_ids or not isinstance(offer_ids, list):
        return {"error": True, "error_message": "商品ID列表不能为空"}
    
    # 确保offer_ids中的元素都是整数
    offer_ids = [int(offer_id) for offer_id in offer_ids if offer_id]
    
    if not offer_ids:
        return {"error": True, "error_message": "处理后的商品ID列表为空"}
    
    # 构建API请求参数
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "offerIds": json.dumps(offer_ids)  # 将商品ID列表转换为JSON字符串
    }
    
    try:
        # 发送API请求
        result = make_api_request(API_PATHS['claim_coupon'], params, user_id=user_id, store_id=store_id)
        
        if isinstance(result, dict) and result.get('error') == True:
            print(f"领取优惠券失败: {result.get('error_message', '未知错误')}")
            return result
        
        print(f"成功领取优惠券，响应结果: {result}")
        return result
    except Exception as e:
        error_msg = f"领取优惠券时发生异常: {str(e)}"
        print(error_msg)
        return {"error": True, "error_message": error_msg}

def bulk_order(user_id=None, store_id=None):
    # 构建查询条件
    query_params = {'订单状态': '等待下单', '供应商': '阿里巴巴网采'}
    
    # 如果提供了store_id，则只处理该store_id的订单
    if store_id:
        query_params['门店ID仓库编码'] = store_id
        
    orders_to_place = query_records('收货单列表', **query_params)
    print(orders_to_place)
    if not orders_to_place:
        print("没有找到需要下单的记录")
        return

    # 按门店ID分组订单
    store_grouped_orders = {}
    for order in orders_to_place:
        order_store_id = order.get('门店ID仓库编码')
        if not order_store_id:
            update_order_status(order, '下单失败', '记录中缺少门店ID')
            continue
            
        if order_store_id not in store_grouped_orders:
            store_grouped_orders[order_store_id] = []
        
        store_grouped_orders[order_store_id].append(order)
    
    # 按门店处理订单
    for order_store_id, store_orders in store_grouped_orders.items():
        # 第一步：获取所有SKU的信息
        sku_info = {}
        all_product_ids = []  # 用于收集所有商品ID，后续领取优惠券使用
        
        for order in store_orders:
            try:
                sku_code = order.get('SKU编码')
                purchase_quantity = order.get('采购量')
                purchase_quantity = validate_and_convert_to_int(purchase_quantity, default=1)
                print(f'正在匹配数据库SKU: {sku_code}')
                
                product_info = query_records('data', SKU编码=sku_code)
                
                if not product_info:
                    print(f"未找到SKU编码为 {sku_code} 的商品信息")
                    update_order_status(order, '下单失败', '未找到对应的商品信息')
                    continue

                product_info = product_info[0]  # 假设只有一条匹配记录
                goods_id = product_info.get('商品ID')
                goods_id = int(float(goods_id)) if goods_id else None
                spec_id = product_info.get('规格ID')
                seller = product_info.get('厂家')
                alibaba_type = product_info.get('采购渠道')
                trans_per_unit = product_info.get('每份单件数')
                trans_per_unit = validate_and_convert_to_int(product_info.get('每份单件数'), default=0)
                purchase_quantity = (purchase_quantity / trans_per_unit) if trans_per_unit != 0 else 1
                purchase_quantity = math.ceil(purchase_quantity)
                min_quality = product_info.get('起购数')
                min_quality = validate_and_convert_to_int(product_info.get('起购数'), default=0)
                purchase_quantity = min_quality if purchase_quantity < min_quality else purchase_quantity
                
                if not goods_id or not spec_id or not seller:
                    update_order_status(order, '下单失败', '商品ID、规格ID或厂家信息缺失')
                    continue
                if '阿里巴巴' not in alibaba_type:
                    update_order_status(order, '下单失败', '下单渠道错误')
                    continue
                
                receive_quantity = purchase_quantity * trans_per_unit  # 计算实际收货数量
                sku_info[sku_code] = {
                    'goods_id': goods_id,
                    'spec_id': None if str(spec_id) == '0' else spec_id,
                    'quantity': purchase_quantity,
                    'seller': seller,
                    'receive_quantity': receive_quantity
                }
                
                # 收集商品ID用于领取优惠券
                if goods_id and goods_id not in all_product_ids:
                    all_product_ids.append(goods_id)
                
            except Exception as e:
                print(f"处理SKU编码为 {sku_code} 的订单时发生错误: {str(e)}")
                update_order_status(order, '下单失败', str(e))
                continue
        
        # 尝试为本次订单中的所有商品领取优惠券
        if all_product_ids:
            try:
                print(f"尝试为商品IDs {all_product_ids} 领取优惠券...")
                coupon_result = claim_optimal_coupon(all_product_ids, store_id=order_store_id)
                if isinstance(coupon_result, dict) and coupon_result.get('error') == True:
                    print(f"领取优惠券失败: {coupon_result.get('error_message')}")
                else:
                    print(f"成功领取优惠券: {coupon_result}")
            except Exception as e:
                print(f"领取优惠券过程中发生错误: {str(e)}")
                # 继续下单流程，不因优惠券领取失败而中断

        # 第二步：按厂家组合订单
        merged_orders = {}
        for sku_code, info in sku_info.items():
            try:
                seller = info['seller']
                if seller not in merged_orders:
                    merged_orders[seller] = []
            
                merged_orders[seller].append({
                    "specId": info['spec_id'],
                    "quantity": info['quantity'],
                    "offerId": info['goods_id']
                })
            except Exception as e:
                print(f"组合订单时发生错误: {str(e)}")
                continue

        # 第三步：下单并更新结果
        for seller, items in merged_orders.items():
            try:
                print(f'正在为厂家 {seller} 下单：{items}')
                # 使用门店ID获取API凭证
                order_result = make_order(items, store_id=order_store_id)
                if order_result.startswith("success:"):
                    order_id = order_result.split(":")[1]
                    for order in store_orders:
                        sku_code = order.get('SKU编码')
                        if sku_code in sku_info and sku_info[sku_code]['seller'] == seller:
                            update_order_status(order, '已下单', order_id=order_id, receive_quantity=sku_info[sku_code]['receive_quantity'], quantity=sku_info[sku_code]['quantity'])
                else:
                    for order in store_orders:
                        sku_code = order.get('SKU编码')
                        if sku_code in sku_info and sku_info[sku_code]['seller'] == seller:
                            update_order_status(order, '下单失败', f'下单失败，{order_result}或请检查阿里巴巴配置')
            except Exception as e:
                print(f"下单时发生错误: {str(e)}")
                for order in store_orders:
                    sku_code = order.get('SKU编码')
                    if sku_code in sku_info and sku_info[sku_code]['seller'] == seller:
                        update_order_status(order, '下单失败', str(e))
                continue
            
    print("下单过程完成")


def update_order_status(order, status, failure_reason=None, order_id=None,receive_quantity=None,quantity=None):
    update_data = {'订单状态': status}
    if status == '下单失败':
        update_data['失败原因'] = failure_reason
    elif status == '已下单':
        update_data['订单编号'] = order_id
        update_data['收货量'] = receive_quantity
        update_data['预留0'] = f'实际下单{quantity}'
    update_records('收货单列表', update_data, id=order['id'])

#读取csv
def process_csv(filename):
    with open(filename, mode='r', encoding='utf-8') as file:
        reader = csv.reader(file)
        rows = list(reader)
    return rows

def check_payment(user_id=None, store_id=None):
    # 构建查询条件
    query_params = {'订单状态': '已下单', '供应商': '阿里巴巴网采'}
    
    # 如果提供了store_id，则只处理该store_id的订单
    if store_id:
        query_params['门店ID仓库编码'] = store_id
    
    # 获取所有数据
    data = query_records('收货单列表', **query_params)
    if not data:
        print("Error fetching data")
        return
    
    # 按门店ID分组订单
    store_grouped_orders = {}
    for row in data:
        order_store_id = row.get('门店ID仓库编码')
        if not order_store_id:
            update_records(
                '收货单列表',
                {'失败原因': '付款失败:记录中缺少门店ID'},
                id=row['id']
            )
            continue
            
        if order_store_id not in store_grouped_orders:
            store_grouped_orders[order_store_id] = []
        
        store_grouped_orders[order_store_id].append(row)
    
    # 按门店处理订单
    for order_store_id, store_orders in store_grouped_orders.items():
        for row in store_orders:
            try:
                sku_code = row['SKU编码']
                order_id = row['订单编号']
                # 获取订单详情（使用门店ID获取API凭证）
                product_details = after_order_preview(order_id, store_id=order_store_id)
                
                # 检查order_id的有效性
                if len(order_id) < 5:
                    error_reason = "订单号错误"
                    update_records(
                        '收货单列表',
                        {'失败原因': f'付款失败:{error_reason}'},
                        订单编号=order_id
                    )
                    continue
                
                print(f'正在付款sku编码: {sku_code}, 订单编号: {order_id}')

                # 检查订单状态
                if product_details == '已支付' or product_details == '订单关闭':
                    print(f'交易状态: {product_details}')
                    update_records(
                        '收货单列表',
                        {'订单状态': product_details,'失败原因':product_details},
                        订单编号=order_id
                    )
                    continue
                elif not product_details:
                    error_reason = '获取商品详情失败'
                    update_records(
                        '收货单列表',
                        {'失败原因': f'付款失败:{error_reason}'},
                        订单编号=order_id
                    )
                    continue
                else:
                    can_pay = False
                    final_price = None
                    
                    for product in product_details:
                        print(product)
                        final_price = float(product['final_price'])
                        specId = product.get('sku_infos_values', 0)
                        quantity = int(product['quantity'])
                        productId = int(product['productId'])
                        
                        # 查询基础数据
                        basedata = query_records('data', 商品ID=productId, 规格ID=specId, 采购渠道='阿里巴巴网采')
                        if basedata:
                            # 确保 basedata 是一个列表，并取第一个元素
                            basedata = basedata[0] if isinstance(basedata, list) and len(basedata) > 0 else basedata
                            historical_cost_str = basedata.get('历史进价', '0')
                            basic_sku = basedata.get('SKU编码', '0')
                            unitTransport = basedata.get('每份单件数', '1')
                            float_unitTransport = float(unitTransport)
                            # final_price = final_price / float_unitTransport if float_unitTransport != 0 else final_price
                            update_records(
                                '收货单列表',
                                {'实际支付': '待同步', '采购单价': final_price},
                                SKU编码=basic_sku,
                                订单编号=order_id
                            )
                        
                            try:
                                historical_cost = float(historical_cost_str) * 1.05
                            except ValueError:
                                historical_cost = 0
                            if historical_cost == 0:
                                can_pay = False
                                error_reason = "数据库没有有效的历史进价信息"
                                update_records(
                                    '收货单列表',
                                    {'失败原因': f'付款失败:{error_reason}'},
                                    订单编号=order_id
                                )
                                break
                            print(f'历史进价{historical_cost}')
                            if final_price <= historical_cost:
                                can_pay = True
                                update_records(
                                    'data',
                                    {
                                        '历史进价': final_price,
                                    },
                                    SKU编码=basic_sku
                                )
                            else:
                                can_pay = False
                                error_reason = "价格高于基础数据"
                                update_records(
                                    '收货单列表',
                                    {
                                        '失败原因': f'付款失败:{error_reason}',
                                        '实际支付': '0',
                                        '采购单价': final_price
                                    },
                                    SKU编码=basic_sku,
                                    订单编号=order_id
                                )
                        else:
                            error_reason = f"基础库未找到商品ID为 {productId} 的商品"
                            update_records(
                                '收货单列表',
                                {'失败原因': f'付款失败:{error_reason}'},
                                订单编号=order_id
                            )

                    if can_pay:
                        # 使用门店凭证进行支付
                        print(f'开始付款订单编号: {order_id}')
                        pay_code = pay_without_pwd(order_id, store_id=order_store_id)
                        time.sleep(1)
                        if pay_code.get('success') == True:
                            update_records(
                                '收货单列表',
                                {'订单状态': '已支付', '失败原因': '付款成功'},
                                订单编号=order_id
                            )
                            print(pay_code)
                        elif pay_code.get('message') == '非待支付订单':
                            update_records(
                                '收货单列表',
                                {'订单状态': '已支付', '失败原因': '付款成功'},
                                订单编号=order_id
                            )
                            print(pay_code)
                        else:
                            error_reason = f'失败支付 {pay_code}'
                            update_records(
                                '收货单列表',
                                {'失败原因': f'付款失败:{error_reason}'},
                                订单编号=order_id
                            )
                        print(pay_code)
            
            except Exception as e:
                error_reason = f"错误: {str(e)}"
                print(f"Error processing row {row['id']}: {error_reason}")
                # 确保 order_id 已被定义
                if 'order_id' in locals():
                    update_records(
                        '收货单列表',
                        {'失败原因': f'付款失败:{error_reason}'},
                        订单编号=order_id
                    )
                else:
                    print("无法更新订单状态，因为 order_id 未定义。")
    
    print("处理完成")
    return 'finish'

def off_order(order_id, user_id=None, store_id=None):
    order_id = int(order_id)
    params = {
        "_aop_timestamp": str(int(time.time() * 1000)),
        "access_token": API_TOKEN,
        "webSite": order_id,
        "tradeID": "general",  # 创建大市场订单
        "cancelReason": "地址错误，重拍！",  # 买家留言（可选）
        # 收货地址信息，假设使用固定的地址信息
        "addressParam": json.dumps(ADDRESS['24客东葛路店']),
        # 下单数据，使用传入的 items 列表直接构建 cargoParamList
        "remark": '地址错误',
        "outOrderId": f"nn_{int(time.time())}",  # 回流订单下游平台（可选）
        "useRedEnvelope": "y"
    }
    result = make_api_request(API_PATHS['off_order'], params, user_id=user_id, store_id=store_id)
    print(result)
    return result

# # Example usage
# if __name__ == "__main__":
#     # # Replace with actual filenames and row numbers
#     csv_filename = '2001_2514.csv'
#     sorted_csv_filename = 'sorted_final.csv'
#     start_row = 2  # Specify starting row (1-based index)
#     end_row = 514  # Specify ending row (1-based index)
    
#     # Check payment eligibility and update CSV for specified rows
#     check_payment(csv_filename, sorted_csv_filename, start_row, end_row)



# if __name__ == "__main__":
    # order_preview(646580127661,'c8217c877f0d3376317e3d9e924dd9ae',1);exit()
#     after_order_preview(2217596595598027671)WW
# 指定起始行和结束行，例如：从第 2001 行到第 2514 行
    # start_row = 11
    # end_row = 364
    # bulk_order(start_row, end_row)
    # make_order(646580127661,'c8217c877f0d3376317e3d9e924dd9ae',1)
    # process_excel()
    # order_ids = get_order_list()
    # for order_id in order_ids:
        # pay_without_pwd(2218160785560027671)
